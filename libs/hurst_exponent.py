import nolds
from pandas import read_csv
from openpyxl import load_workbook

workbook = load_workbook(filename="../data/chaos_data/results_presentation.xlsx")
worksheets = workbook.sheetnames

for sheet in worksheets:
    # Activate worksheet to write dataframe
    active = workbook[sheet]

    # load dataset
    series = read_csv('../data/datasets/' + sheet + '_final_week.csv',
                      header=0, index_col=0, parse_dates=True, squeeze=True)
    x = series.values
    x = x.astype('float32')
    # x = np.fromiter(series.values, dtype="float32")

    hurst_exponent = nolds.hurst_rs(x)
    print("Sample entropy: " + str(hurst_exponent))

    # columns = list(range(2, 5))
    # emb_dim = -1
    #
    # for column in columns:
    #     if column == 2:
    #         emb_dim = 3
    #     elif column == 3:
    #         emb_dim = 5
    #     elif column == 4:
    #         emb_dim = 4
    #     # Do the calculation and put it on a specific cell
    #     sample_entropy = nolds.sampen(x, emb_dim=emb_dim, tolerance=None)
    #     print("Sample entropy: " + str(sample_entropy))
    #     active.cell(row=27, column=column).value = sample_entropy

# Save workbook to write
# workbook.save("../data/chaos_data/results_presentation.xlsx")
workbook.close()
