import nolds
from pandas import read_csv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
from openpyxl.formatting.rule import CellIsRule

# lm = nolds.logistic_map(0.1, 1000, r=4)
# x = np.fromiter(lm, dtype="float32")
# l = max(nolds.lyap_e(x))
# k = nolds.lyap_r(x)
# print(l)
# print(k)

# ======================================================================================================================
# Example using Eckman's method
# ======================================================================================================================
# emb_dim_ext = 3
# matrix_dim_ext = 3
# min_nb_ext = min(2 * matrix_dim_ext, matrix_dim_ext + 4)
# e = max(nolds.lyap_e(x, emb_dim=emb_dim_ext, matrix_dim=matrix_dim_ext, min_nb=min_nb_ext))
# print("Eckmann: " + str(e))

# ======================================================================================================================
# Main calculations using Rosenstein's method
# ======================================================================================================================
workbook = load_workbook(filename="../data/chaos_data/results_presentation.xlsx")
worksheets = workbook.sheetnames
green_fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

for sheet in worksheets:
    # Activate worksheet to write dataframe
    active = workbook[sheet]
    active.conditional_formatting.add('B3:M22', CellIsRule(operator='greaterThan', formula=[0.2], fill=green_fill))

    # load dataset
    series = read_csv('../data/datasets/' + sheet + '_final_week.csv',
                      header=0, index_col=0, parse_dates=True, squeeze=True)
    x = series.values
    x = x.astype('float32')
    # x = np.fromiter(series.values, dtype="float32")

    rows = list(range(3, 7)) + list(range(11, 15)) + list(range(19, 23))
    columns = list(range(2, 14))

    emb_dim = [3, 5, 7]
    lag = [1, 2, 3, 4]
    min_neighbors = [2, 3, 4, 5]
    trajectory_len = [6, 7, 8]

    for row in rows:
        if row == 3:
            trajectory_len = 6
            lag = 1
        elif row == 4:
            trajectory_len = 6
            lag = 2
        elif row == 5:
            trajectory_len = 6
            lag = 3
        elif row == 6:
            trajectory_len = 6
            lag = 4
        elif row == 11:
            trajectory_len = 7
            lag = 1
        elif row == 12:
            trajectory_len = 7
            lag = 2
        elif row == 13:
            trajectory_len = 7
            lag = 3
        elif row == 14:
            trajectory_len = 7
            lag = 4
        elif row == 19:
            trajectory_len = 8
            lag = 1
        elif row == 20:
            trajectory_len = 8
            lag = 2
        elif row == 21:
            trajectory_len = 8
            lag = 3
        elif row == 22:
            trajectory_len = 8
            lag = 4
        for column in columns:
            if column == 2:
                emb_dim = 3
                min_neighbors = 2
            elif column == 3:
                emb_dim = 3
                min_neighbors = 3
            elif column == 4:
                emb_dim = 3
                min_neighbors = 4
            elif column == 5:
                emb_dim = 3
                min_neighbors = 5
            elif column == 6:
                emb_dim = 5
                min_neighbors = 2
            elif column == 7:
                emb_dim = 5
                min_neighbors = 3
            elif column == 8:
                emb_dim = 5
                min_neighbors = 4
            elif column == 9:
                emb_dim = 5
                min_neighbors = 5
            elif column == 10:
                emb_dim = 7
                min_neighbors = 2
            elif column == 11:
                emb_dim = 7
                min_neighbors = 3
            elif column == 12:
                emb_dim = 7
                min_neighbors = 4
            elif column == 13:
                emb_dim = 7
                min_neighbors = 5
            # Do the calculation and put it on a specific cell
            r = nolds.lyap_r(x, emb_dim=emb_dim, lag=lag, min_tsep=None, min_neighbors=min_neighbors,
                             trajectory_len=trajectory_len)
            active.cell(row=row, column=column).value = r

# Save workbook to write
workbook.save("../data/chaos_data/results_presentation.xlsx")
workbook.close()
