import nolds
from pandas import read_csv
from openpyxl import load_workbook

workbook = load_workbook(filename="../data/chaos_data/results_presentation.xlsx")
worksheets = workbook.sheetnames

for sheet in worksheets:
    # Activate worksheet to write dataframe
    active = workbook[sheet]

    # load dataset
    series = read_csv('../data/datasets/' + sheet + '_final_week.csv',
                      header=0, index_col=0, parse_dates=True, squeeze=True)
    x = series.values
    x = x.astype('float32')
    # x = np.fromiter(series.values, dtype="float32")

    hurst_exponent = nolds.hurst_rs(x)
    print("Hurst exponent: " + str(hurst_exponent))

workbook.close()
